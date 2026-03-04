from fastapi import APIRouter, Depends, HTTPException, Query, Request, Response, UploadFile, File, Form
from fastapi.responses import StreamingResponse, RedirectResponse
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, func, and_, or_
from sqlalchemy.orm import selectinload
from datetime import datetime, timezone
from typing import Optional, List
import uuid
import logging

logger = logging.getLogger(__name__)

from app.core.database import get_db
from app.core.security import (
    decrypt_text, encrypt_text,
    decrypt_case_content, encrypt_case_content,
    decrypt_comment_content, encrypt_comment_content,
)
from app.models import (
    Case, CaseAttachment, CaseComment, AuditLog, AuditAction, CaseStatus,
    CasePriority, CaseCategory, User, UserRole
)
from app.schemas.cases import (
    CaseResponse, CaseListResponse, CaseDetailResponse,
    AssignCaseRequest, ChangeStatusRequest, AddCommentRequest,
    CaseCommentResponse
)
from app.api.v1.auth import get_current_user, require_investigator_or_above, require_admin
from app.services.notification import send_reporter_message
from app.services.storage import get_file_content, get_download_url, delete_file
from app.core.config import settings

router = APIRouter(prefix="/cases", tags=["cases"])
logger = logging.getLogger(__name__)

# ─── Status holat mashinasi ──────────────────────────────────────────────────
ALLOWED_TRANSITIONS: dict[CaseStatus, list[CaseStatus]] = {
    CaseStatus.NEW:         [CaseStatus.IN_PROGRESS, CaseStatus.REJECTED, CaseStatus.NEEDS_INFO],
    CaseStatus.IN_PROGRESS: [CaseStatus.COMPLETED, CaseStatus.REJECTED, CaseStatus.NEEDS_INFO, CaseStatus.NEW],
    CaseStatus.NEEDS_INFO:  [CaseStatus.IN_PROGRESS, CaseStatus.REJECTED],
    CaseStatus.COMPLETED:   [],   # faqat admin → archived
    CaseStatus.REJECTED:    [],   # faqat admin → archived
    CaseStatus.ARCHIVED:    [],
}
# Admin faqat yakunlangan/rad etilgan → arxivga o'tkaza oladi
ADMIN_ONLY_TRANSITIONS: dict[CaseStatus, list[CaseStatus]] = {
    CaseStatus.COMPLETED: [CaseStatus.ARCHIVED],
    CaseStatus.REJECTED:  [CaseStatus.ARCHIVED],
}


def decrypt_case(case: Case) -> dict:
    base = {
        "id": str(case.id),
        "external_id": case.external_id,
        "category": case.category,
        "priority": case.priority,
        "status": case.status,
        "title": case.title,
        "is_anonymous": case.is_anonymous,
        "created_at": case.created_at,
        "updated_at": case.updated_at,
        "closed_at": case.closed_at,
        "due_at": case.due_at,
        "assigned_to": str(case.assigned_to) if case.assigned_to else None,
        "attachments_count": len(case.attachments) if case.attachments else 0,
        "jira_ticket_id": case.jira_ticket_id,
        "jira_ticket_url": case.jira_ticket_url,
    }
    return base


@router.get("/export")
async def export_cases(
    format: str = Query("xlsx", pattern="^(xlsx|pdf)$"),
    status: Optional[CaseStatus] = None,
    category: Optional[CaseCategory] = None,
    priority: Optional[CasePriority] = None,
    from_date: Optional[str] = Query(None, description="ISO 8601, masalan 2026-01-01"),
    to_date:   Optional[str] = Query(None, description="ISO 8601, masalan 2026-12-31"),
    current_user: User = Depends(require_investigator_or_above),
    db: AsyncSession = Depends(get_db),
):
    """Murojaatlarni Excel (.xlsx) yoki PDF (.pdf) formatida eksport qilish."""
    from datetime import date as dt_date

    # ── Filtr ──────────────────────────────────────────────────────────────
    qfilters = []
    if status:
        qfilters.append(Case.status == status)
    if category:
        qfilters.append(Case.category == category)
    if priority:
        qfilters.append(Case.priority == priority)
    if from_date:
        try:
            fd = datetime.fromisoformat(from_date).replace(tzinfo=timezone.utc)
            qfilters.append(Case.created_at >= fd)
        except ValueError:
            pass
    if to_date:
        try:
            td = datetime.fromisoformat(to_date).replace(tzinfo=timezone.utc)
            # to_date kuni oxirigacha
            td = td.replace(hour=23, minute=59, second=59)
            qfilters.append(Case.created_at <= td)
        except ValueError:
            pass

    query = (
        select(Case)
        .options(selectinload(Case.assignee))
        .order_by(Case.created_at.desc())
    )
    if qfilters:
        query = query.where(and_(*qfilters))

    result = await db.execute(query)
    cases = result.scalars().all()

    # ── Label lug'atlari ───────────────────────────────────────────────────
    cat_labels = {
        "corruption": "Korrupsiya / Pora",
        "conflict_of_interest": "Manfaatlar to'qnashuvi",
        "fraud": "Firibgarlik / O'g'irlik",
        "safety": "Xavfsizlik buzilishi",
        "discrimination": "Kamsitish / Bezovtalik",
        "procurement": "Tender buzilishi",
        "other": "Boshqa",
    }
    status_labels = {
        "new": "Yangi",
        "in_progress": "Ko'rib chiqilmoqda",
        "needs_info": "Ma'lumot kerak",
        "completed": "Yakunlandi",
        "rejected": "Rad etildi",
        "archived": "Arxivlandi",
    }
    priority_labels = {
        "critical": "Kritik",
        "high": "Yuqori",
        "medium": "O'rta",
        "low": "Past",
    }

    export_date = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")
    file_date   = datetime.now(timezone.utc).strftime("%Y-%m-%d")

    # ── Filter info matni ──────────────────────────────────────────────────
    filter_parts = []
    if status:   filter_parts.append(f"Holat: {status_labels.get(status.value, status.value)}")
    if category: filter_parts.append(f"Kategoriya: {cat_labels.get(category.value, category.value)}")
    if priority: filter_parts.append(f"Ustuvorlik: {priority_labels.get(priority.value, priority.value)}")
    if from_date: filter_parts.append(f"Dan: {from_date}")
    if to_date:   filter_parts.append(f"Gacha: {to_date}")
    filter_info = " | ".join(filter_parts) if filter_parts else "Barcha murojaatlar"

    # ── Row ma'lumotlarini tayyorlash ──────────────────────────────────────
    rows = []
    for c in cases:
        try:
            desc = decrypt_case_content(c.description_encrypted)[:500]
        except Exception:
            desc = "[Shifrni ochib bo'lmadi]"
        assignee_name = ""
        if c.assignee:
            assignee_name = c.assignee.full_name or c.assignee.username
        rows.append({
            "id":       c.external_id,
            "category": cat_labels.get(c.category.value if hasattr(c.category, "value") else c.category, str(c.category)),
            "status":   status_labels.get(c.status.value if hasattr(c.status, "value") else c.status, str(c.status)),
            "priority": priority_labels.get(c.priority.value if hasattr(c.priority, "value") else c.priority, str(c.priority)),
            "anon":     "Ha" if c.is_anonymous else "Yo'q",
            "assignee": assignee_name,
            "created":  c.created_at.strftime("%d.%m.%Y %H:%M") if c.created_at else "",
            "closed":   c.closed_at.strftime("%d.%m.%Y %H:%M") if c.closed_at else "",
            "desc":     desc,
            "status_raw": c.status.value if hasattr(c.status, "value") else str(c.status),
        })

    headers_row = ["Murojaat ID", "Tavsif", "Kategoriya", "Holat", "Ustuvorlik",
                   "Anonimlik", "Ijrochi", "Yaratilgan", "Yopilgan"]

    # ══════════════════════════════════════════════════════════════════════
    # EXCEL
    # ══════════════════════════════════════════════════════════════════════
    if format == "xlsx":
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        import io

        wb = Workbook()
        ws = wb.active
        ws.title = "Murojaatlar"

        # ── 1-qator: sarlavha ──────────────────────────────────────────────
        ws.merge_cells("A1:I1")
        ws["A1"] = f"IntegrityBot — Eksport hisoboti  |  {export_date}"
        ws["A1"].font = Font(bold=True, size=13, color="FFFFFF")
        ws["A1"].fill = PatternFill("solid", fgColor="1E3A5F")
        ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 24

        # ── 2-qator: filter info ───────────────────────────────────────────
        ws.merge_cells("A2:I2")
        ws["A2"] = f"Filtr: {filter_info}  |  Jami: {len(rows)} ta murojaat"
        ws["A2"].font = Font(italic=True, size=10, color="444444")
        ws["A2"].fill = PatternFill("solid", fgColor="D9E1F2")
        ws["A2"].alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[2].height = 18

        # ── 3-qator: ustun sarlavhalari ────────────────────────────────────
        header_fill = PatternFill("solid", fgColor="2E75B6")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        thin = Side(style="thin", color="AAAAAA")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        for col_idx, col_name in enumerate(headers_row, start=1):
            cell = ws.cell(row=3, column=col_idx, value=col_name)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border
        ws.row_dimensions[3].height = 22

        # ── Rang xaritalari (holat bo'yicha) ───────────────────────────────
        row_fills = {
            "completed":   PatternFill("solid", fgColor="CCFFCC"),
            "rejected":    PatternFill("solid", fgColor="FFCCCC"),
            "in_progress": PatternFill("solid", fgColor="FFFFCC"),
            "needs_info":  PatternFill("solid", fgColor="FFE4B5"),
            "archived":    PatternFill("solid", fgColor="E0E0E0"),
        }

        # ── Ma'lumot qatorlari ─────────────────────────────────────────────
        for r_idx, row in enumerate(rows, start=4):
            fill = row_fills.get(row["status_raw"], None)
            vals = [row["id"], row["desc"], row["category"], row["status"], row["priority"],
                    row["anon"], row["assignee"], row["created"], row["closed"]]
            for c_idx, val in enumerate(vals, start=1):
                cell = ws.cell(row=r_idx, column=c_idx, value=val)
                cell.border = border
                cell.alignment = Alignment(vertical="top", wrap_text=(c_idx == 2))
                if fill:
                    cell.fill = fill
                if c_idx == 1:
                    cell.font = Font(bold=True, color="1F497D")

        # ── Ustun kengliklarini auto-fit ───────────────────────────────────
        col_widths = [18, 60, 22, 18, 12, 10, 20, 18, 18]
        for i, w in enumerate(col_widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w

        # ── Freeze top rows ────────────────────────────────────────────────
        ws.freeze_panes = "A4"

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        filename = f"integrity_report_{file_date}.xlsx"
        return StreamingResponse(
            buf,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'},
        )

    # ══════════════════════════════════════════════════════════════════════
    # PDF
    # ══════════════════════════════════════════════════════════════════════
    else:
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib import colors
        from reportlab.lib.units import mm
        from reportlab.platypus import (
            SimpleDocTemplate, Table, TableStyle, Paragraph,
            Spacer, HRFlowable
        )
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.enums import TA_CENTER, TA_LEFT
        import io

        buf = io.BytesIO()
        doc = SimpleDocTemplate(
            buf,
            pagesize=landscape(A4),
            leftMargin=15*mm, rightMargin=15*mm,
            topMargin=18*mm, bottomMargin=18*mm,
        )

        styles = getSampleStyleSheet()
        title_style = ParagraphStyle("title", parent=styles["Title"],
            fontSize=16, spaceAfter=4, textColor=colors.HexColor("#1E3A5F"))
        sub_style = ParagraphStyle("sub", parent=styles["Normal"],
            fontSize=9, textColor=colors.grey, spaceAfter=10)
        cell_style = ParagraphStyle("cell", parent=styles["Normal"],
            fontSize=7.5, leading=10)
        header_style = ParagraphStyle("hdr", parent=styles["Normal"],
            fontSize=8, textColor=colors.white, fontName="Helvetica-Bold")

        story = []

        # Sarlavha
        story.append(Paragraph("IntegrityBot — Eksport hisoboti", title_style))
        story.append(Paragraph(f"Sana: {export_date}  |  Filtr: {filter_info}  |  Jami: {len(rows)} ta", sub_style))
        story.append(HRFlowable(width="100%", thickness=1, color=colors.HexColor("#2E75B6"), spaceAfter=8))

        # Jadval sarlavhalari
        col_widths_pdf = [38*mm, 65*mm, 35*mm, 28*mm, 20*mm, 14*mm, 28*mm, 24*mm, 24*mm]
        header_cells = [Paragraph(h, header_style) for h in headers_row]
        table_data = [header_cells]

        status_colors_pdf = {
            "completed":   colors.HexColor("#CCFFCC"),
            "rejected":    colors.HexColor("#FFCCCC"),
            "in_progress": colors.HexColor("#FFFFCC"),
            "needs_info":  colors.HexColor("#FFE4B5"),
            "archived":    colors.HexColor("#E0E0E0"),
        }

        row_bg_cmds = []
        for r_idx, row in enumerate(rows, start=1):
            cells = [
                Paragraph(row["id"], cell_style),
                Paragraph(row["desc"][:300], cell_style),
                Paragraph(row["category"], cell_style),
                Paragraph(row["status"], cell_style),
                Paragraph(row["priority"], cell_style),
                Paragraph(row["anon"], cell_style),
                Paragraph(row["assignee"], cell_style),
                Paragraph(row["created"], cell_style),
                Paragraph(row["closed"], cell_style),
            ]
            table_data.append(cells)
            bg = status_colors_pdf.get(row["status_raw"])
            if bg:
                row_bg_cmds.append(("BACKGROUND", (0, r_idx), (-1, r_idx), bg))

        tbl = Table(table_data, colWidths=col_widths_pdf, repeatRows=1)
        tbl_style = TableStyle([
            # Sarlavha
            ("BACKGROUND",  (0, 0), (-1, 0), colors.HexColor("#2E75B6")),
            ("TEXTCOLOR",   (0, 0), (-1, 0), colors.white),
            ("FONTNAME",    (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE",    (0, 0), (-1, 0), 8),
            ("ALIGN",       (0, 0), (-1, 0), "CENTER"),
            ("VALIGN",      (0, 0), (-1, -1), "TOP"),
            ("FONTSIZE",    (0, 1), (-1, -1), 7.5),
            ("GRID",        (0, 0), (-1, -1), 0.4, colors.HexColor("#BBBBBB")),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F5F8FF")]),
            ("LEFTPADDING",  (0, 0), (-1, -1), 4),
            ("RIGHTPADDING", (0, 0), (-1, -1), 4),
            ("TOPPADDING",   (0, 0), (-1, -1), 3),
            ("BOTTOMPADDING",(0, 0), (-1, -1), 3),
        ] + row_bg_cmds)
        tbl.setStyle(tbl_style)
        story.append(tbl)

        # Altbilgi funksiyasi
        def add_page_number(canvas_obj, doc_obj):
            canvas_obj.saveState()
            canvas_obj.setFont("Helvetica", 8)
            canvas_obj.setFillColor(colors.grey)
            page_text = f"Sahifa {doc_obj.page}  —  IntegrityBot eksport  —  {export_date}"
            canvas_obj.drawCentredString(landscape(A4)[0] / 2, 10*mm, page_text)
            canvas_obj.restoreState()

        doc.build(story, onFirstPage=add_page_number, onLaterPages=add_page_number)
        buf.seek(0)

        filename = f"integrity_report_{file_date}.pdf"
        return StreamingResponse(
            buf,
            media_type="application/pdf",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'},
        )


@router.get("", response_model=dict)
async def list_cases(
    page: int = Query(1, ge=1),
    per_page: int = Query(20, ge=1, le=100),
    status: Optional[CaseStatus] = None,
    category: Optional[CaseCategory] = None,
    priority: Optional[CasePriority] = None,
    assigned_to: Optional[str] = None,
    search: Optional[str] = None,
    current_user: User = Depends(require_investigator_or_above),
    db: AsyncSession = Depends(get_db),
):
    filters = []
    if status:
        filters.append(Case.status == status)
    if category:
        filters.append(Case.category == category)
    if priority:
        filters.append(Case.priority == priority)
    if assigned_to:
        filters.append(Case.assigned_to == uuid.UUID(assigned_to))

    query = select(Case).options(selectinload(Case.attachments))
    if filters:
        query = query.where(and_(*filters))
    query = query.order_by(Case.created_at.desc())

    total_result = await db.execute(select(func.count(Case.id)).where(and_(*filters) if filters else True))
    total = total_result.scalar()

    offset = (page - 1) * per_page
    result = await db.execute(query.offset(offset).limit(per_page))
    cases = result.scalars().all()

    return {
        "items": [decrypt_case(c) for c in cases],
        "total": total,
        "page": page,
        "per_page": per_page,
        "pages": (total + per_page - 1) // per_page,
    }


@router.get("/stats")
async def get_stats(
    period: Optional[str] = Query(None, pattern="^(today|week|month|year)$"),
    from_date: Optional[str] = Query(None, description="ISO date, e.g. 2026-01-01"),
    to_date:   Optional[str] = Query(None, description="ISO date, e.g. 2026-12-31"),
    current_user: User = Depends(require_investigator_or_above),
    db: AsyncSession = Depends(get_db),
):
    from sqlalchemy import extract, cast, Date as SADate, Integer as SAInteger, text
    now = datetime.now(timezone.utc)

    # ── Sana oralig'ini aniqlash ──────────────────────────────────────────
    from_dt: Optional[datetime] = None
    to_dt:   Optional[datetime] = None

    if from_date or to_date:
        # from_date/to_date ustunlik qiladi
        if from_date:
            try:
                from_dt = datetime.fromisoformat(from_date).replace(
                    hour=0, minute=0, second=0, microsecond=0, tzinfo=timezone.utc)
            except ValueError:
                pass
        if to_date:
            try:
                to_dt = datetime.fromisoformat(to_date).replace(
                    hour=23, minute=59, second=59, microsecond=999999, tzinfo=timezone.utc)
            except ValueError:
                pass
    elif period:
        if period == "today":
            from_dt = now.replace(hour=0, minute=0, second=0, microsecond=0)
            to_dt   = now
        elif period == "week":
            # Haftaning dushanbasi
            monday = now - __import__('datetime').timedelta(days=now.weekday())
            from_dt = monday.replace(hour=0, minute=0, second=0, microsecond=0)
            to_dt   = now
        elif period == "month":
            from_dt = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
            to_dt   = now
        elif period == "year":
            from_dt = now.replace(month=1, day=1, hour=0, minute=0, second=0, microsecond=0)
            to_dt   = now

    # ── WHERE shartini qurish ─────────────────────────────────────────────
    def date_filter():
        conds = []
        if from_dt:
            conds.append(Case.created_at >= from_dt)
        if to_dt:
            conds.append(Case.created_at <= to_dt)
        return and_(*conds) if conds else True

    # ── 1. Total ──────────────────────────────────────────────────────────
    total_r = await db.execute(
        select(func.count(Case.id)).where(date_filter())
    )
    total = total_r.scalar() or 0

    # ── 2. By status (1 query) ────────────────────────────────────────────
    status_r = await db.execute(
        select(Case.status, func.count(Case.id))
        .where(date_filter())
        .group_by(Case.status)
    )
    status_counts = {s.value: 0 for s in CaseStatus}
    for row in status_r:
        key = row[0].value if hasattr(row[0], "value") else str(row[0])
        status_counts[key] = row[1]

    # ── 3. By category (1 query) ──────────────────────────────────────────
    cat_r = await db.execute(
        select(Case.category, func.count(Case.id))
        .where(date_filter())
        .group_by(Case.category)
    )
    cat_counts = {c.value: 0 for c in CaseCategory}
    for row in cat_r:
        key = row[0].value if hasattr(row[0], "value") else str(row[0])
        cat_counts[key] = row[1]

    # ── 4. By priority (1 query) ──────────────────────────────────────────
    pri_r = await db.execute(
        select(Case.priority, func.count(Case.id))
        .where(date_filter())
        .group_by(Case.priority)
    )
    pri_counts = {p.value: 0 for p in CasePriority}
    for row in pri_r:
        key = row[0].value if hasattr(row[0], "value") else str(row[0])
        pri_counts[key] = row[1]

    # ── 5. Trend (1 query, smart grouping) ───────────────────────────────
    # Oraliq uzunligi bo'yicha guruhlash strategiyasini belgilash
    if from_dt and to_dt:
        delta_days = (to_dt - from_dt).days
    elif from_dt:
        delta_days = (now - from_dt).days
    else:
        delta_days = 180  # default: oy bo'yicha

    monthly = []

    if delta_days < 7:
        # Kun bo'yicha guruhlash
        fd = from_dt or (now - __import__('datetime').timedelta(days=6)).replace(
            hour=0, minute=0, second=0, microsecond=0)
        td = to_dt or now
        import datetime as dt_mod
        current = fd.replace(hour=0, minute=0, second=0, microsecond=0)
        while current.date() <= td.date():
            next_day = current + dt_mod.timedelta(days=1)
            day_conds = [Case.created_at >= current, Case.created_at < next_day]
            if from_dt: day_conds.append(Case.created_at >= from_dt)
            if to_dt:   day_conds.append(Case.created_at <= to_dt)
            r = await db.execute(
                select(func.count(Case.id)).where(and_(*day_conds))
            )
            monthly.append({
                "month": current.strftime("%d %b"),
                "count": r.scalar() or 0,
            })
            current = next_day

    elif delta_days <= 60:
        # Hafta bo'yicha guruhlash
        import datetime as dt_mod
        fd = from_dt or (now - dt_mod.timedelta(weeks=8)).replace(
            hour=0, minute=0, second=0, microsecond=0)
        td = to_dt or now
        # Haftaning boshiga to'g'rila
        week_start = fd - dt_mod.timedelta(days=fd.weekday())
        week_start = week_start.replace(hour=0, minute=0, second=0, microsecond=0)
        while week_start <= td:
            week_end = week_start + dt_mod.timedelta(days=7)
            r = await db.execute(
                select(func.count(Case.id)).where(
                    and_(Case.created_at >= week_start, Case.created_at < week_end,
                         Case.created_at >= fd, Case.created_at <= td)
                )
            )
            monthly.append({
                "month": week_start.strftime("%d %b"),
                "count": r.scalar() or 0,
            })
            week_start = week_end

    else:
        # Oy bo'yicha guruhlash — so'nggi 12 oy (yoki oraliq)
        import datetime as dt_mod
        if from_dt:
            start_year, start_month = from_dt.year, from_dt.month
        else:
            start_month = (now.month - 11) % 12 or 12
            start_year  = now.year - (1 if now.month <= 11 else 0)
        end_year, end_month = (to_dt or now).year, (to_dt or now).month
        y, m = start_year, start_month
        while (y, m) <= (end_year, end_month):
            month_start = datetime(y, m, 1, tzinfo=timezone.utc)
            if m == 12:
                month_end = datetime(y + 1, 1, 1, tzinfo=timezone.utc)
            else:
                month_end = datetime(y, m + 1, 1, tzinfo=timezone.utc)
            conds = [Case.created_at >= month_start, Case.created_at < month_end]
            if from_dt: conds.append(Case.created_at >= from_dt)
            if to_dt:   conds.append(Case.created_at <= to_dt)
            r = await db.execute(select(func.count(Case.id)).where(and_(*conds)))
            monthly.append({
                "month": f"{y}-{m:02d}",
                "count": r.scalar() or 0,
            })
            if m == 12:
                y += 1; m = 1
            else:
                m += 1

    # Trend label ni qaytarish uchun meta
    trend_label = "oy"
    if delta_days < 7:
        trend_label = "kun"
    elif delta_days <= 60:
        trend_label = "hafta"

    return {
        "total": total,
        "by_status": status_counts,
        "by_category": cat_counts,
        "by_priority": pri_counts,
        "monthly_trend": monthly,
        "trend_label": trend_label,
        "filter": {
            "period": period,
            "from_date": from_dt.date().isoformat() if from_dt else None,
            "to_date": to_dt.date().isoformat() if to_dt else None,
        },
    }



@router.get("/{case_id}")
async def get_case(
    case_id: str,
    request: Request,
    current_user: User = Depends(require_investigator_or_above),
    db: AsyncSession = Depends(get_db),
):
    result = await db.execute(
        select(Case)
        .options(
            selectinload(Case.attachments),
            selectinload(Case.comments).selectinload(CaseComment.author),
            selectinload(Case.assignee),
        )
        .where(Case.external_id == case_id)
    )
    case = result.scalar_one_or_none()
    if not case:
        raise HTTPException(status_code=404, detail="Case not found")

    # Audit log
    db.add(AuditLog(
        user_id=current_user.id,
        case_id=case.id,
        action=AuditAction.CASE_VIEW,
        ip_address=request.client.host if request.client else None,
    ))
    await db.commit()

    base = decrypt_case(case)
    base["description"] = decrypt_case_content(case.description_encrypted)

    # reporter_ip SAQLANMAYDI — anonimlik kafolati (ISO 37001, O'zbekiston shaxsiy ma'lumotlar qonuni)
    # Audit log'dagi IP faqat ADMIN harakatlarini qayd etadi, reporter IP'si emas

    base["comments"] = []
    for c in sorted(case.comments, key=lambda x: x.created_at):
        base["comments"].append({
            "id": str(c.id),
            "content": decrypt_comment_content(c.content_encrypted),
            "is_from_reporter": c.is_from_reporter,
            "is_internal": c.is_internal,
            "author": c.author.full_name if c.author else ("Reporter" if c.is_from_reporter else "System"),
            "created_at": c.created_at,
        })
    base["attachments"] = [
        {
            "id": str(a.id),
            "filename": a.original_filename,
            "mime_type": a.mime_type,
            "size_bytes": a.size_bytes,
            "uploaded_at": a.uploaded_at,
            "uploaded_by_admin": getattr(a, "uploaded_by_admin", False),
        }
        for a in case.attachments
    ]
    if case.assignee:
        base["assignee_name"] = case.assignee.full_name or case.assignee.username

    return base


@router.get("/{case_id}/export")
async def export_case_pdf(
    case_id: str,
    request: Request,
    current_user: User = Depends(require_investigator_or_above),
    db: AsyncSession = Depends(get_db),
):
    """Bitta murojaat kartochkasini PDF sifatida eksport qilish."""
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.units import mm
    from reportlab.platypus import (
        SimpleDocTemplate, Table, TableStyle, Paragraph,
        Spacer, HRFlowable
    )
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER, TA_LEFT
    import io

    result = await db.execute(
        select(Case)
        .options(selectinload(Case.attachments), selectinload(Case.comments).selectinload(CaseComment.author), selectinload(Case.assignee))
        .where(Case.external_id == case_id)
    )
    case = result.scalar_one_or_none()
    if not case:
        raise HTTPException(status_code=404, detail="Case not found")

    cat_labels = {
        "corruption": "Korrupsiya / Pora", "conflict_of_interest": "Manfaatlar to'qnashuvi",
        "fraud": "Firibgarlik", "safety": "Xavfsizlik buzilishi",
        "discrimination": "Kamsitish", "procurement": "Tender buzilishi", "other": "Boshqa",
    }
    status_labels = {
        "new": "Yangi", "in_progress": "Ko'rib chiqilmoqda", "needs_info": "Ma'lumot kerak",
        "completed": "Yakunlandi", "rejected": "Rad etildi", "archived": "Arxivlandi",
    }
    priority_labels = {"critical": "Kritik", "high": "Yuqori", "medium": "O'rta", "low": "Past"}

    def v(enum_val):
        return enum_val.value if hasattr(enum_val, "value") else str(enum_val)

    try:
        description = decrypt_case_content(case.description_encrypted)
    except Exception:
        description = "[Shifrni ochib bo'lmadi]"

    export_dt = datetime.now(timezone.utc).strftime("%d.%m.%Y %H:%M UTC")

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4,
        leftMargin=20*mm, rightMargin=20*mm, topMargin=20*mm, bottomMargin=20*mm)

    styles = getSampleStyleSheet()
    h1 = ParagraphStyle("h1", fontSize=16, fontName="Helvetica-Bold",
        textColor=colors.HexColor("#1E3A5F"), spaceAfter=4)
    h2 = ParagraphStyle("h2", fontSize=11, fontName="Helvetica-Bold",
        textColor=colors.HexColor("#2E75B6"), spaceBefore=12, spaceAfter=4)
    normal = ParagraphStyle("n", fontSize=9, leading=13)
    label_s = ParagraphStyle("lbl", fontSize=8, textColor=colors.grey)
    value_s = ParagraphStyle("val", fontSize=9, fontName="Helvetica-Bold")

    story = []

    # Sarlavha
    story.append(Paragraph(f"Murojaat kartochkasi: {case.external_id}", h1))
    story.append(Paragraph(f"Eksport sanasi: {export_dt}", label_s))
    story.append(HRFlowable(width="100%", thickness=1.5, color=colors.HexColor("#2E75B6"), spaceAfter=10))

    # Asosiy ma'lumotlar jadvali
    info_data = [
        ["Maydon", "Qiymat"],
        ["Murojaat ID", case.external_id],
        ["Kategoriya", cat_labels.get(v(case.category), v(case.category))],
        ["Holat", status_labels.get(v(case.status), v(case.status))],
        ["Ustuvorlik", priority_labels.get(v(case.priority), v(case.priority))],
        ["Anonimlik", "Ha" if case.is_anonymous else "Yo'q"],
        ["Ijrochi", (case.assignee.full_name or case.assignee.username) if case.assignee else "Tayinlanmagan"],
        ["Yaratilgan", case.created_at.strftime("%d.%m.%Y %H:%M") if case.created_at else "—"],
        ["Yopilgan", case.closed_at.strftime("%d.%m.%Y %H:%M") if case.closed_at else "—"],
        ["Muddat", case.due_at.strftime("%d.%m.%Y") if case.due_at else "—"],
    ]
    info_tbl = Table(info_data, colWidths=[50*mm, 120*mm])
    info_tbl.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2E75B6")),
        ("TEXTCOLOR",  (0, 0), (-1, 0), colors.white),
        ("FONTNAME",   (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE",   (0, 0), (-1, -1), 9),
        ("BACKGROUND", (0, 1), (0, -1), colors.HexColor("#F0F4FA")),
        ("FONTNAME",   (0, 1), (0, -1), "Helvetica-Bold"),
        ("GRID",       (0, 0), (-1, -1), 0.5, colors.HexColor("#CCCCCC")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F5F8FF")]),
        ("LEFTPADDING",  (0, 0), (-1, -1), 6),
        ("TOPPADDING",   (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING",(0, 0), (-1, -1), 4),
    ]))
    story.append(info_tbl)

    # Tavsif
    story.append(Paragraph("Murojaat mazmuni", h2))
    story.append(Paragraph(description.replace("\n", "<br/>"), normal))

    # Chat tarixi
    comments = sorted(case.comments, key=lambda x: x.created_at)
    if comments:
        story.append(Paragraph("Chat tarixi", h2))
        chat_data = [["Vaqt", "Kim", "Xabar", "Tur"]]
        for cm in comments:
            try:
                content = decrypt_comment_content(cm.content_encrypted)[:300]
            except Exception:
                content = "[Shifrni ochib bo'lmadi]"
            who = "Reporter" if cm.is_from_reporter else (
                cm.author.full_name if cm.author else "Admin")
            kind = "Ichki" if cm.is_internal else "Tashqi"
            ts = cm.created_at.strftime("%d.%m %H:%M") if cm.created_at else ""
            chat_data.append([ts, who, content, kind])
        chat_tbl = Table(chat_data, colWidths=[22*mm, 28*mm, 100*mm, 16*mm])
        chat_tbl.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2E75B6")),
            ("TEXTCOLOR",  (0, 0), (-1, 0), colors.white),
            ("FONTNAME",   (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE",   (0, 0), (-1, -1), 8),
            ("GRID",       (0, 0), (-1, -1), 0.4, colors.HexColor("#CCCCCC")),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F5F8FF")]),
            ("VALIGN",     (0, 0), (-1, -1), "TOP"),
            ("LEFTPADDING",  (0, 0), (-1, -1), 4),
            ("TOPPADDING",   (0, 0), (-1, -1), 3),
            ("BOTTOMPADDING",(0, 0), (-1, -1), 3),
        ]))
        story.append(chat_tbl)

    # Fayllar ro'yxati
    if case.attachments:
        story.append(Paragraph("Biriktirилgan fayllar", h2))
        att_data = [["Fayl nomi", "Tur", "Hajm", "Sana", "Yuklagan"]]
        for a in case.attachments:
            size = f"{a.size_bytes // 1024} KB" if a.size_bytes else "—"
            uploaded = "Admin" if getattr(a, "uploaded_by_admin", False) else "Reporter"
            att_data.append([
                a.original_filename,
                a.mime_type,
                size,
                a.uploaded_at.strftime("%d.%m.%Y") if a.uploaded_at else "—",
                uploaded,
            ])
        att_tbl = Table(att_data, colWidths=[60*mm, 35*mm, 20*mm, 25*mm, 25*mm])
        att_tbl.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2E75B6")),
            ("TEXTCOLOR",  (0, 0), (-1, 0), colors.white),
            ("FONTNAME",   (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE",   (0, 0), (-1, -1), 8),
            ("GRID",       (0, 0), (-1, -1), 0.4, colors.HexColor("#CCCCCC")),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F5F8FF")]),
            ("LEFTPADDING",  (0, 0), (-1, -1), 4),
            ("TOPPADDING",   (0, 0), (-1, -1), 3),
        ]))
        story.append(att_tbl)

    def footer(canvas_obj, doc_obj):
        canvas_obj.saveState()
        canvas_obj.setFont("Helvetica", 7.5)
        canvas_obj.setFillColor(colors.grey)
        canvas_obj.drawCentredString(A4[0] / 2, 10*mm,
            f"Sahifa {doc_obj.page}  —  IntegrityBot  —  {export_dt}  —  Maxfiy")
        canvas_obj.restoreState()

    doc.build(story, onFirstPage=footer, onLaterPages=footer)
    buf.seek(0)

    filename = f"case_{case_id}_{datetime.now(timezone.utc).strftime('%Y-%m-%d')}.pdf"
    return StreamingResponse(
        buf,
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.post("/{case_id}/assign")
async def assign_case(
    case_id: str,
    body: AssignCaseRequest,
    request: Request,
    current_user: User = Depends(require_investigator_or_above),
    db: AsyncSession = Depends(get_db),
):
    result = await db.execute(select(Case).where(Case.external_id == case_id))
    case = result.scalar_one_or_none()
    if not case:
        raise HTTPException(status_code=404, detail="Case not found")

    case.assigned_to = uuid.UUID(body.user_id) if body.user_id else None
    db.add(AuditLog(
        user_id=current_user.id,
        case_id=case.id,
        action=AuditAction.CASE_ASSIGN,
        payload={"assigned_to": body.user_id},
        ip_address=request.client.host if request.client else None,
    ))
    await db.commit()

    # ── Ijrochiga Telegram bildirishnoma ──────────────────────────────────
    if body.user_id:
        try:
            assignee_result = await db.execute(
                select(User).where(User.id == uuid.UUID(body.user_id))
            )
            assignee = assignee_result.scalar_one_or_none()
            if assignee and assignee.telegram_chat_id:
                from app.api.v1.telegram import get_bot_app
                from app.services.notification import notify_assignee
                bot_app = get_bot_app()
                due_date = case.due_at.strftime("%d.%m.%Y") if case.due_at else "Belgilanmagan"
                await notify_assignee(
                    bot=bot_app.bot,
                    case_id=case.external_id,
                    category=case.category,
                    priority=case.priority,
                    due_date=due_date,
                    assignee_telegram_id=assignee.telegram_chat_id,
                )
        except Exception as e:
            logger.warning(f"Assignee notification failed for {case_id}: {e}")

    return {"message": "Case assigned"}


@router.post("/{case_id}/status")
async def change_status(
    case_id: str,
    body: ChangeStatusRequest,
    request: Request,
    current_user: User = Depends(require_investigator_or_above),
    db: AsyncSession = Depends(get_db),
):
    result = await db.execute(select(Case).where(Case.external_id == case_id))
    case = result.scalar_one_or_none()
    if not case:
        raise HTTPException(status_code=404, detail="Case not found")

    old_status = case.status
    new_status = body.status

    # ── Holat mashinasi validatsiyasi ──────────────────────────────────────
    is_admin = current_user.role == UserRole.ADMIN
    allowed = list(ALLOWED_TRANSITIONS.get(old_status, []))
    if is_admin:
        allowed += ADMIN_ONLY_TRANSITIONS.get(old_status, [])

    if new_status not in allowed:
        allowed_vals = [s.value for s in allowed] or []
        raise HTTPException(
            status_code=400,
            detail=f"'{old_status.value}' → '{new_status.value}' o'tish ruxsat etilmagan. "
                   f"Ruxsat etilganlar: {allowed_vals if allowed_vals else 'yoq'}",
        )

    # ── Rad etishda izoh majburiy ──────────────────────────────────────────
    if new_status == CaseStatus.REJECTED and not (body.reason or "").strip():
        raise HTTPException(
            status_code=422,
            detail="Rad etish uchun sabab (reason) majburiy",
        )

    # ── Holat yangilash ────────────────────────────────────────────────────
    case.status = new_status
    if new_status in (CaseStatus.COMPLETED, CaseStatus.REJECTED):
        case.closed_at = datetime.now(timezone.utc)
    elif new_status == CaseStatus.ARCHIVED:
        pass  # closed_at saqlanadi

    # ── Izohni comment jadvaliga yoz (agar reason berilgan bo'lsa) ─────────
    reason_text = (body.reason or "").strip()
    if reason_text:
        status_label = new_status.value.replace("_", " ").title()
        comment = CaseComment(
            case_id=case.id,
            author_id=current_user.id,
            is_from_reporter=False,
            is_internal=True,
            content_encrypted=encrypt_comment_content(
                f"[Holat o'zgartirildi: {old_status.value} → {new_status.value}]\n\n{reason_text}"
            ),
        )
        db.add(comment)

    db.add(AuditLog(
        user_id=current_user.id,
        case_id=case.id,
        action=AuditAction.CASE_UPDATE,
        payload={
            "old_status": old_status.value,
            "new_status": new_status.value,
            "reason": reason_text or None,
        },
        ip_address=request.client.host if request.client else None,
    ))
    await db.commit()

    # ── Jira tiket yangilash ───────────────────────────────────────────────
    if getattr(case, "jira_ticket_id", None):
        try:
            from app.services.jira_integration import ticket_service
            await ticket_service.update_ticket_on_case_status_change(
                ticket_id=case.jira_ticket_id,
                new_status=new_status.value,
                case_id=case_id,
            )
        except Exception as e:
            logger.warning(f"Tiket yangilashda xato ({case_id}): {e}")

    return {"message": "Status updated", "status": new_status.value}


@router.post("/{case_id}/comment")
async def add_comment(
    case_id: str,
    body: AddCommentRequest,
    request: Request,
    current_user: User = Depends(require_investigator_or_above),
    db: AsyncSession = Depends(get_db),
):
    result = await db.execute(select(Case).where(Case.external_id == case_id))
    case = result.scalar_one_or_none()
    if not case:
        raise HTTPException(status_code=404, detail="Case not found")

    comment = CaseComment(
        case_id=case.id,
        author_id=current_user.id,
        is_from_reporter=False,
        is_internal=body.is_internal,
        content_encrypted=encrypt_comment_content(body.content),
    )
    db.add(comment)

    # Ichki eslatma bo'lmasa — har doim reporterga Telegram orqali yuboramiz
    # (anonim bo'lsa ham chat_id saqlanadi, shaxsiy ma'lumot emas)
    if not body.is_internal and case.telegram_chat_id:
        try:
            from app.api.v1.telegram import get_bot_app
            bot_app = get_bot_app()
            await send_reporter_message(bot_app.bot, case.telegram_chat_id, case.external_id, body.content)
        except Exception as e:
            logger.warning(f"Could not send Telegram reply for {case.external_id}: {e}")

    db.add(AuditLog(
        user_id=current_user.id,
        case_id=case.id,
        action=AuditAction.CASE_COMMENT,
        payload={"is_internal": body.is_internal},
        ip_address=request.client.host if request.client else None,
    ))
    await db.commit()
    return {"message": "Comment added"}


@router.get("/{case_id}/attachments/{attachment_id}")
async def download_attachment(
    case_id: str,
    attachment_id: str,
    request: Request,
    current_user: User = Depends(require_investigator_or_above),
    db: AsyncSession = Depends(get_db),
):
    result = await db.execute(select(Case).where(Case.external_id == case_id))
    case = result.scalar_one_or_none()
    if not case:
        raise HTTPException(status_code=404, detail="Case not found")

    att_result = await db.execute(
        select(CaseAttachment).where(
            CaseAttachment.id == uuid.UUID(attachment_id),
            CaseAttachment.case_id == case.id,
        )
    )
    attachment = att_result.scalar_one_or_none()
    if not attachment:
        raise HTTPException(status_code=404, detail="Attachment not found")

    db.add(AuditLog(
        user_id=current_user.id,
        case_id=case.id,
        action=AuditAction.ATTACHMENT_DOWNLOAD,
        payload={"attachment_id": str(attachment.id), "filename": attachment.original_filename},
        ip_address=request.client.host if request.client else None,
    ))
    await db.commit()

    if settings.STORAGE_TYPE == "s3":
        presigned_url = await get_download_url(attachment.storage_path, expires_in=300)
        return RedirectResponse(url=presigned_url, status_code=302)

    try:
        content = await get_file_content(attachment.storage_path)
    except FileNotFoundError:
        raise HTTPException(status_code=404, detail="File not found on disk")

    mime = attachment.mime_type or "application/octet-stream"
    is_viewable = (
        mime.startswith("image/") or
        mime.startswith("video/") or
        mime.startswith("audio/") or
        mime == "application/pdf"
    )
    disposition = "inline" if is_viewable else f'attachment; filename="{attachment.original_filename}"'
    total_size = len(content)

    # Video/Audio uchun Range request (streaming) qo'llab-quvvatlash
    range_header = request.headers.get("range")
    if range_header and (mime.startswith("video/") or mime.startswith("audio/")):
        try:
            range_val = range_header.strip().replace("bytes=", "")
            start_str, end_str = range_val.split("-")
            start = int(start_str)
            end = int(end_str) if end_str else total_size - 1
            end = min(end, total_size - 1)
            chunk = content[start:end + 1]
            return Response(
                content=chunk,
                status_code=206,
                media_type=mime,
                headers={
                    "Content-Range": f"bytes {start}-{end}/{total_size}",
                    "Accept-Ranges": "bytes",
                    "Content-Length": str(len(chunk)),
                    "Content-Disposition": disposition,
                    "Cache-Control": "private, max-age=3600",
                },
            )
        except Exception:
            pass  # Range parse xatosi — oddiy response qaytaramiz

    return Response(
        content=content,
        status_code=200,
        media_type=mime,
        headers={
            "Content-Disposition": disposition,
            "Content-Length": str(total_size),
            "Accept-Ranges": "bytes",
            "Cache-Control": "private, max-age=3600",
        },
    )


@router.post("/{case_id}/send-file")
async def send_file_to_reporter(
    case_id: str,
    request: Request,
    file: UploadFile = File(...),
    caption: str = Form(default=""),
    is_internal: bool = Form(default=False),
    current_user: User = Depends(require_investigator_or_above),
    db: AsyncSession = Depends(get_db),
):
    """Admin reporter ga fayl/rasm/video/audio yuboradi (Telegram orqali)"""
    from app.services.storage import (
        save_uploaded_file, ALLOWED_MIME_TYPES, BLOCKED_EXTENSIONS, MAX_FILE_SIZE_BYTES
    )
    from app.services.notification import send_reporter_file
    from app.services.storage import sanitize_filename
    import os
    import hashlib

    result = await db.execute(select(Case).where(Case.external_id == case_id))
    case = result.scalar_one_or_none()
    if not case:
        raise HTTPException(status_code=404, detail="Case not found")

    # Fayl validatsiya
    file_data = await file.read()
    if len(file_data) > MAX_FILE_SIZE_BYTES:
        raise HTTPException(400, detail=f"Fayl hajmi {settings.MAX_FILE_SIZE_MB}MB dan oshmasligi kerak")

    ext = os.path.splitext(file.filename or "")[1].lower()
    if ext in BLOCKED_EXTENSIONS:
        raise HTTPException(400, detail=f"'{ext}' kengaytmali fayllar yuklash taqiqlangan")

    # MIME aniqlash: brauzer noto'g'ri MIME yuborsa — kengaytmadan aniqlaymiz
    import mimetypes
    mime = file.content_type or "application/octet-stream"
    # Brauzer octet-stream yuborsa yoki video/quicktime ni mp4 sifatida yuborsa — kengaytmadan aniqlaymiz
    if mime in ("application/octet-stream", "application/x-www-form-urlencoded"):
        guessed, _ = mimetypes.guess_type(file.filename or "")
        if guessed:
            mime = guessed

    # Video kengaytmalarini qo'shimcha tekshirish (brauzer har xil MIME yuborishi mumkin)
    video_exts = {".mp4", ".avi", ".mov", ".mkv", ".webm", ".flv"}
    audio_exts = {".mp3", ".wav", ".ogg", ".aac", ".flac", ".m4a", ".opus", ".oga"}
    image_exts = {".jpg", ".jpeg", ".png", ".gif", ".webp", ".bmp"}

    if ext in video_exts and not mime.startswith("video/"):
        mime = "video/mp4" if ext == ".mp4" else f"video/{ext.lstrip('.')}"
    elif ext in audio_exts and not mime.startswith("audio/"):
        mime = "audio/mpeg" if ext == ".mp3" else f"audio/{ext.lstrip('.')}"
    elif ext in image_exts and not mime.startswith("image/"):
        mime = f"image/{'jpeg' if ext in ('.jpg', '.jpeg') else ext.lstrip('.')}"

    # Hali ham ruxsat etilmagan MIME bo'lsa — rad etamiz (lekin video/audio/image uchun o'tkazamiz)
    if (mime not in ALLOWED_MIME_TYPES and
            not mime.startswith("video/") and
            not mime.startswith("audio/") and
            not mime.startswith("image/")):
        raise HTTPException(400, detail=f"Bu fayl turi ruxsat etilmagan: {mime}")

    # Faylni storage ga saqlash
    safe_filename = file.filename or f"file_{uuid.uuid4().hex[:8]}"
    storage_path = await save_uploaded_file(
        file_data=file_data,
        original_filename=safe_filename,
        mime_type=mime,
        case_id=str(case.id),
    )

    # DB ga attachment saqlaymiz (admin yuklagan deb belgilanadi)
    safe_name = sanitize_filename(safe_filename)
    checksum = hashlib.sha256(file_data).hexdigest()
    attachment = CaseAttachment(
        case_id=case.id,
        storage_path=storage_path,
        filename=safe_name,
        original_filename=safe_filename,
        mime_type=mime,
        size_bytes=len(file_data),
        checksum=checksum,
        uploaded_by_admin=True,
    )
    db.add(attachment)
    await db.flush()  # attachment.id ni olish uchun

    # Comment yozamiz (chat ichida fayl ko'rinsin)
    comment_text = f"📎 Fayl yuborildi: {safe_filename}"
    if caption:
        comment_text += f"\n{caption}"
    comment = CaseComment(
        case_id=case.id,
        author_id=current_user.id,
        is_from_reporter=False,
        is_internal=is_internal,
        content_encrypted=encrypt_comment_content(comment_text),
    )
    db.add(comment)

    # Ichki eslatma bo'lmasa — reporter ga Telegram orqali fayl yuboramiz
    tg_sent = False
    if not is_internal and case.telegram_chat_id:
        try:
            from app.api.v1.telegram import get_bot_app
            bot_app = get_bot_app()
            await send_reporter_file(
                bot=bot_app.bot,
                telegram_chat_id=case.telegram_chat_id,
                case_id=case.external_id,
                file_data=file_data,
                filename=safe_filename,
                mime_type=mime,
                caption=caption,
            )
            tg_sent = True
        except Exception as e:
            logger.warning(f"Telegram file send failed for {case_id}: {e}")

    db.add(AuditLog(
        user_id=current_user.id,
        case_id=case.id,
        action=AuditAction.CASE_COMMENT,
        payload={
            "action": "admin_sent_file",
            "filename": safe_filename,
            "mime_type": mime,
            "is_internal": is_internal,
            "telegram_sent": tg_sent,
        },
        ip_address=request.client.host if request.client else None,
    ))
    await db.commit()

    return {
        "message": "Fayl yuborildi",
        "attachment_id": str(attachment.id),
        "telegram_sent": tg_sent,
        "filename": safe_filename,
    }


