from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, func, and_
from sqlalchemy.orm import selectinload
from typing import Optional
from datetime import date, datetime, timezone, timedelta
import logging
import uuid
import io

from app.core.database import get_db
from app.models import AuditLog, User, Case
from app.api.v1.auth import require_admin

router = APIRouter(prefix="/audit", tags=["audit"])
logger = logging.getLogger(__name__)


def _build_audit_filters(
    action: Optional[str],
    user_id: Optional[str],
    from_date: Optional[date],
    to_date: Optional[date],
    case_id: Optional[str],
):
    """Audit log filtrlarini yaratish."""
    filters = []
    need_join_case = False

    if action:
        filters.append(AuditLog.action == action)
    if user_id:
        filters.append(AuditLog.user_id == uuid.UUID(user_id))
    if from_date:
        filters.append(
            AuditLog.created_at >= datetime(from_date.year, from_date.month, from_date.day, tzinfo=timezone.utc)
        )
    if to_date:
        to_dt = datetime(to_date.year, to_date.month, to_date.day, tzinfo=timezone.utc) + timedelta(days=1)
        filters.append(AuditLog.created_at < to_dt)
    if case_id:
        filters.append(Case.external_id == case_id)
        need_join_case = True

    return filters, need_join_case


def _serialize_log(log):
    return {
        "id": str(log.id),
        "action": log.action.value if hasattr(log.action, "value") else str(log.action),
        "entity_type": log.entity_type,
        "entity_id": log.entity_id,
        "ip_address": log.ip_address,
        "payload": log.payload,
        "created_at": log.created_at.isoformat() if log.created_at else None,
        "user": {
            "id": str(log.user.id),
            "username": log.user.username,
            "full_name": log.user.full_name,
            "role": log.user.role.value if hasattr(log.user.role, "value") else str(log.user.role),
        } if log.user else None,
        "case_id": str(log.case_id) if log.case_id else None,
    }


@router.get("", response_model=dict)
async def list_audit_logs(
    page: int = Query(1, ge=1),
    per_page: int = Query(50, ge=1, le=200),
    action: Optional[str] = None,
    user_id: Optional[str] = None,
    from_date: Optional[date] = None,
    to_date: Optional[date] = None,
    case_id: Optional[str] = None,
    current_user: User = Depends(require_admin),
    db: AsyncSession = Depends(get_db),
):
    filters, need_join_case = _build_audit_filters(action, user_id, from_date, to_date, case_id)

    base_query = select(AuditLog).options(selectinload(AuditLog.user))
    count_query = select(func.count(AuditLog.id))

    if need_join_case:
        base_query = base_query.join(Case, AuditLog.case_id == Case.id)
        count_query = count_query.join(Case, AuditLog.case_id == Case.id)

    if filters:
        base_query = base_query.where(and_(*filters))
        count_query = count_query.where(and_(*filters))

    total_result = await db.execute(count_query)
    total = total_result.scalar()

    offset = (page - 1) * per_page
    result = await db.execute(
        base_query.order_by(AuditLog.created_at.desc()).offset(offset).limit(per_page)
    )
    logs = result.scalars().all()

    return {
        "items": [_serialize_log(log) for log in logs],
        "total": total,
        "page": page,
        "per_page": per_page,
        "pages": max(1, -(-total // per_page)),
    }


@router.get("/export")
async def export_audit_logs(
    format: str = Query("xlsx", pattern="^xlsx$"),
    action: Optional[str] = None,
    user_id: Optional[str] = None,
    from_date: Optional[date] = None,
    to_date: Optional[date] = None,
    case_id: Optional[str] = None,
    current_user: User = Depends(require_admin),
    db: AsyncSession = Depends(get_db),
):
    """Audit loglarni Excel ga eksport (faqat admin uchun)."""
    filters, need_join_case = _build_audit_filters(action, user_id, from_date, to_date, case_id)

    query = select(AuditLog).options(selectinload(AuditLog.user))
    if need_join_case:
        query = query.join(Case, AuditLog.case_id == Case.id)
    if filters:
        query = query.where(and_(*filters))

    result = await db.execute(query.order_by(AuditLog.created_at.desc()))
    logs = result.scalars().all()

    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = Workbook()
    ws = wb.active
    ws.title = "Audit Log"

    headers = ["Sana", "Foydalanuvchi", "Amal", "Ob'ekt turi", "Ob'ekt ID", "IP manzil", "Izoh"]
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for row_idx, log in enumerate(logs, 2):
        ws.cell(row=row_idx, column=1, value=log.created_at.strftime("%Y-%m-%d %H:%M:%S") if log.created_at else "")
        ws.cell(row=row_idx, column=2, value=(log.user.full_name or log.user.username) if log.user else "Tizim")
        ws.cell(row=row_idx, column=3, value=log.action.value if hasattr(log.action, "value") else str(log.action))
        ws.cell(row=row_idx, column=4, value=log.entity_type or "")
        ws.cell(row=row_idx, column=5, value=log.entity_id or "")
        ws.cell(row=row_idx, column=6, value=log.ip_address or "")
        ws.cell(row=row_idx, column=7, value=str(log.payload) if log.payload else "")

    # Ustun kengliklarini sozlash
    for col_idx, width in enumerate([20, 25, 18, 15, 15, 18, 40], 1):
        ws.column_dimensions[chr(64 + col_idx)].width = width

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"audit_log_{datetime.now(timezone.utc).strftime('%Y%m%d_%H%M%S')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/actions", response_model=list[str])
async def list_audit_actions(
    current_user: User = Depends(require_admin),
):
    """Barcha mavjud audit action'lar ro'yxati"""
    from app.models import AuditAction
    return [a.value for a in AuditAction]


@router.get("/retention/stats")
async def get_retention_stats(
    current_user: User = Depends(require_admin),
    db: AsyncSession = Depends(get_db),
):
    """
    Data retention statistikasi — nechta yozuv muddati yaqinlashmoqda.
    Admin dashboard uchun.
    """
    from app.services.retention import get_retention_stats as _get_stats
    return await _get_stats(db)


@router.post("/retention/run")
async def run_retention_now(
    current_user: User = Depends(require_admin),
):
    """
    Data retention'ni qo'lda ishga tushirish (admin only).
    Odatda har kecha 02:00 UTC da avtomatik ishlaydi.
    """
    from app.services.retention import run_retention
    logger.info(f"Manual retention triggered by {current_user.username}")
    stats = await run_retention()
    return {
        "message": "Data retention muvaffaqiyatli bajarildi",
        "stats": stats,
    }

